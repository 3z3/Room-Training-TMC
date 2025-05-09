import openpyxl

wb = openpyxl.load_workbook('..\\TMCroomlist.xlsx')    #path to the excel file being read, update depending on how you organize your files

ws = wb.active

memory_to_rooms = {}
memory_to_areas = {}

for i in range(2,ws.max_row+1):
    area_cell = ws.cell(row = i, column = 2)
    if area_cell.value is None or area_cell.fill.bgColor.rgb != 'FFFFFFFF':    #cells that are not white are bugged rooms or rooms that are not used in the game
        pass
    else:
        if area_cell.value not in memory_to_rooms:
            memory_to_areas[ws.cell(row = i, column = 1).value] = area_cell.value
            
        room_cell = ws.cell(row = i, column = 4)
        if room_cell.value is None or room_cell.fill.bgColor.rgb != 'FFFFFFFF':
            pass
        else:
            memory_to_rooms[ws.cell(row = i, column = 3).value + ws.cell(row = i, column = 1).value] = room_cell.value

mtrstring = 'memory_to_rooms = {'
mtastring = 'memory_to_areas = {'

#block needed since some areas have the same name on the excel file but different memory addresses
area_individual_names = {}

for memory,name in memory_to_areas.items():
    area_individual_names[name] = area_individual_names.get(name,0) + 1
    if area_individual_names[name] > 1:
        memory_to_areas[memory] = name + " " + string(area_individual_names[name])    #count, eg. "Dark Hyrule Castle 2"

#translates the memory_to_rooms dictionary into Lua
for key,value in memory_to_rooms.items():
    mtrstring += ' ["' + key + '"] = "' + value + '",'
mtrstring[:-1] += ' }\n'    #getting rid of the last comma at the end

#translates the memory_to_areas dictionary into Lua
for key,value in memory_to_areas.items():
    mtastring += ' ["' + key + '"] = "' + value + '",'
mtastring[:-1] += ' }'

with open("roomnames.Lua", "w") as f:
  f.writelines([mtrstring,mtastring])
